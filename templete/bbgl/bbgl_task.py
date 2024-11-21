#!/usr/bin/env python3
import urllib.parse
import urllib.request
import traceback
import pymysql
import sys
import warnings
import datetime
import json
import ssl
import requests
import logging
import openpyxl
import os,zipfile
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime,timedelta

def send_mail465(p_to_user,p_to_cc,p_title,p_content):
    try:
        sys_sender = '190343@lifeat.cn'  # 系统账户
        sys_pwd = 'R86hyfjobMBYR76h'  # 系统账户密码
        to_user = p_to_user.split(",")
        to_cc = p_to_cc.split(",")
        msg = MIMEText(p_content,'html','utf-8')
        msg["Subject"] = p_title
        msg['From'] = formataddr(["技术服务部", sys_sender])
        msg["To"] = ",".join(to_user)
        msg['Cc'] = ",".join(to_cc)
        server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465, timeout=60)
        server.set_debuglevel(0)
        server.login(sys_sender, sys_pwd)
        server.sendmail(sys_sender,
                        [formataddr(["", u]) for u in to_user] +
                        [formataddr(["", u]) for u in to_cc],
                        msg.as_string())
        server.quit()
    except smtplib.SMTPException as e:
        traceback.print_exc()

class SendMail(object):
    '''
    https://cloud.tencent.com/developer/article/1598257
    '''
    def __init__(self,sender,Cc,title,content):
        self.sender = sender #.split(',')  #发送地址
        self.Cc = Cc #.split(',')  # 抄送地址
        self.title = title  # 标题
        self.content = content  # 发送内容
        self.sys_sender = '190343@lifeat.cn'  # 系统账户
        self.sys_pwd = 'R86hyfjobMBYR76h'  # 系统账户密码

    def send(self,file_list):
        """
        发送邮件
        :param file_list: 附件文件列表
        :return: bool
        """
        try:
            # 创建一个带附件的实例
            msg = MIMEMultipart()
            # 发件人格式
            msg['From'] = formataddr(["技术服务部", self.sys_sender])
            # 收件人格式
            print('a=',[formataddr(["", u]) for u in self.sender.split(',')])
            msg['To'] = self.sender

            # 抄送人
            msg['Cc'] = self.Cc

            # 邮件主题
            msg['Subject'] = self.title

            # 邮件正文内容
            msg.attach(MIMEText(self.content, 'html', 'utf-8'))

            # 多个附件
            for file_name in file_list:
                print("file_name",file_name)
                # 构造附件
                xlsxpart = MIMEApplication(open(file_name, 'rb').read())
                # filename表示邮件中显示的附件名
                xlsxpart.add_header('Content-Disposition','attachment',filename = '%s'%file_name)
                msg.attach(xlsxpart)

            # SMTP服务器
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465,timeout=60)
            # 登录账户
            server.login(self.sys_sender, self.sys_pwd)
            # 发送邮件
            #server.sendmail(self.sys_sender, [self.sender, ], msg.as_string())
            server.sendmail(self.sys_sender,
                            [formataddr(["", u]) for u in self.sender.split(',')]+
                            [formataddr(["", u]) for u in self.Cc.split(',')],
                            msg.as_string())
            # 退出账户
            server.quit()
            return True
        except Exception as e:
            traceback.print_exc()
            print(e)
            return False

def print_dict(config):
    print('-'.ljust(85,'-'))
    print(' '.ljust(3,' ')+"name".ljust(20,' ')+'value')
    print('-'.ljust(85,'-'))
    for key in config:
      print(' '.ljust(3,' ')+key.ljust(20,' ')+'=',config[key])
    print('-'.ljust(85,'-'))

def get_ds_mysql(ip,port,service ,user,password):
    try:
        conn = pymysql.connect(host=ip, port=int(port), user=user, passwd=password, db=service, charset='utf8',connect_timeout=3)
        return conn
    except  Exception as e:
        print('get_ds_mysql exceptiion:' + traceback.format_exc())
        return None

def aes_decrypt(p_password,p_key):
    values = {
        'password': p_password,
        'key':p_key
    }
    url = 'http://$$API_SERVER$$/read_db_decrypt'
    context = ssl._create_unverified_context()
    data = urllib.parse.urlencode(values).encode(encoding='UTF-8')
    req = urllib.request.Request(url, data=data)
    res = urllib.request.urlopen(req, context=context)
    res = json.loads(res.read())
    if res['code'] == 200:
        print('接口read_db_decrypt 调用成功!')
        config = res['msg']
        return config
    else:
        print('接口read_db_decrypt 调用失败!,{0}'.format(res['msg']))
        sys.exit(0)

def send_mail(config):
    pass

def get_config_from_db(tag):
    url = 'http://$$API_SERVER$$/read_config_bbgl'
    res = None
    try:
        res = requests.post(url, data={'tag': tag}, timeout=30).json()
    except:
        traceback.print_exc()

    if res is not None :
        if res['code'] == 200:
            print(res['msg'])
            config = res['msg']
            db_ip = config['db_ip']
            db_port = config['db_port']
            db_service = config['db_service']
            db_user = config['db_user']
            db_pass = aes_decrypt(config['db_pass'], db_user)
            config['db_string'] = 'msyql://' + db_ip + ':' + db_port + '/' + db_service
            config['db_mysql'] = get_ds_mysql(db_ip, db_port, db_service, db_user, db_pass)
            return config
        else:
           logging.info(res['msg'])
           print(res['msg'])
           return None
    else:
      return None

def init(tag):
    config = get_config_from_db(tag)
    return config

def current_rq():
    year =str(datetime.now().year)
    month=str(datetime.now().month).rjust(2,'0')
    day  =str(datetime.now().day).rjust(2,'0')
    return year+month+day

def current_time():
    now_time = datetime.now()
    time1_str = datetime.strftime(now_time, '%Y-%m-%d %H:%M:%S')
    return time1_str

def get_tjrq_range(cfg):
    cr = cfg['db_mysql'].cursor()
    cr.execute('select {},{}'.format(cfg['tjrq_begin'],cfg['tjrq_end']))
    rs=cr.fetchone()
    return '{} 0:0:0 - {} 23:59:59'.format(rs[0],rs[1])

def exp_data(cfg,p_header,p_data):
    row_data  = 1
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0,title=cfg['bbmc'])
    file_path = cfg['script_path'] + '/bbtj'
    if not os.path.exists(file_path):
       os.mkdir(file_path)
    os.system('cd {0}'.format(file_path))
    file_name = cfg['script_path'] + '/bbtj/exp_bbtj_{}_{}.xlsx'.format(cfg['bbdm'],current_rq())
    arch_name = 'exp_bbtj_{}_{}.xlsx'.format(cfg['bbdm'],current_rq())

    # write header
    for k in range(len(p_header)):
        ws.cell(column = k+1,row=row_data,value = p_header[k]['header_name'])

    # write body
    row_data = row_data + 1
    for i in p_data:
        for j in range(len(i)):
            if i[j] is None:
               ws.cell(row=row_data, column=j+1,value='')
            else:
               ws.cell(row=row_data, column=j+1,value = str(i[j]))
        row_data = row_data + 1
    wb.save(file_name)
    print("{0} export complete!".format(file_name))
    zip_file = 'bbtj_{}_{}.zip'.format(cfg['bbdm'],current_rq())

    if os.path.exists(zip_file):
        os.system('rm -f {0}'.format(zip_file))

    z = zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED, allowZip64=True)
    z.write(file_name, arcname=arch_name)
    z.close()
    os.system('rm -f {0}'.format(file_name))
    return zip_file

def run_report(cfg):
    headers = cfg['headers']
    preprocess = cfg['preproccess']
    param = {
         'bbrq_begin':cfg['tjrq_begin'],
         'bbrq_end':"concat({},' 23:59:59')".format(cfg['tjrq_end'])
    }
    cr = cfg['db_mysql'].cursor()

    if preprocess != []:
        # preprocess param replace
        for s in preprocess:
            s['replace_statement'] = s['statement']
            for key, value in param.items():
                s['replace_statement'] = s['replace_statement'].replace('\'$$' + key + '$$\'', value)
                logging.info('preprocess : '+s['replace_statement'])
                cr.execute(s['replace_statement'])
    cfg['db_mysql'].commit()

    # process query
    if cfg['statement'] is not None and cfg['statement'] != '':
        cfg['replace_statement'] = cfg['statement']
        for key, value in param.items():
            cfg['replace_statement'] = cfg['replace_statement'].replace('\'$$' + key + '$$\'', value)

        # exec query
        logging.info(cfg['replace_statement'])
        cr.execute(cfg['replace_statement'])
        des = [c[0]  for c in cr.description]
        logging.info(des)
        res = cr.fetchall()

        logging.info('write data to excel...')
        zip_file = exp_data(cfg,headers,res)

        # send mail
        sender = cfg['receiver']
        Cc = cfg['cc']

        title = '{}({})'.format(cfg['bbmc'],current_rq())
        content = '''各位领导：<br><br>
    &nbsp;&nbsp;&nbsp;【{}】报表数据已生成完成。<br>
    &nbsp;&nbsp;&nbsp;&nbsp;统计时间 ：{}<br>
    &nbsp;&nbsp;&nbsp;&nbsp;统计范围 ：{}<br><br>         
    详见附件,请查收。
            '''.format(cfg['bbmc'],current_time(),get_tjrq_range(cfg))

        # add attentment
        file_list = [zip_file]
        ret = SendMail(sender, Cc, title, content).send(file_list)
        logging.info('mail send success')
    else:
        # send mail
        sender = cfg['receiver']
        Cc = cfg['cc']
        title = '{}({})'.format(cfg['bbmc'], current_rq())
        content = '''各位领导：<br><br>
           &nbsp;&nbsp;&nbsp;【{}】报表数据已生成完成。<br>
           &nbsp;&nbsp;&nbsp;&nbsp;统计时间 ：{}<br>
           &nbsp;&nbsp;&nbsp;&nbsp;统计范围 ：{}<br><br>         
          请知晓。
'''.format(cfg['bbmc'], current_time(), get_tjrq_range(cfg))
        send_mail465(sender, Cc, title, content)
        logging.info('mail send success')

def main():
    tag = ""
    warnings.filterwarnings("ignore")
    for p in range(len(sys.argv)):
        if sys.argv[p] == "-tag":
            tag = sys.argv[p + 1]

    # init logger
    logging.basicConfig(filename='/tmp/{}.{}.log'.format(tag, datetime.now().strftime("%Y-%m-%d")),
                        format='[%(asctime)s-%(levelname)s:%(message)s]',
                        level=logging.INFO, filemode='a', datefmt='%Y-%m-%d %I:%M:%S')

    config=init(tag)
    if config:
        print_dict(config)
        run_report(config)



if __name__ == "__main__":
     main()
